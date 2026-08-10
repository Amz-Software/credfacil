from django import forms
from django.contrib import admin, messages
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.html import format_html
from simple_history.admin import SimpleHistoryAdmin
from .models import *

class AdminBase(admin.ModelAdmin):
    list_display = ('loja', 'criado_em', 'modificado_em')
    readonly_fields = ('criado_em', 'modificado_em')

    def save_model(self, request, obj, form, change):
        obj.save(user=request.user)
        super().save_model(request, obj, form, change)


class HistoryAdminBase(SimpleHistoryAdmin):
    list_display = ('loja', 'criado_em', 'modificado_em')
    readonly_fields = ('criado_em', 'modificado_em')

    def save_model(self, request, obj, form, change):
        obj.save(user=request.user)
        super().save_model(request, obj, form, change)

class ProdutoVendaInline(admin.TabularInline):
    model = ProdutoVenda
    extra = 1

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.criado_por = request.user
        obj.modificado_por = request.user
        super().save_model(request, obj, form, change)

class PagamentoInline(admin.TabularInline):
    model = Pagamento
    extra = 1

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.criado_por = request.user
        obj.modificado_por = request.user
        super().save_model(request, obj, form, change)

@admin.register(Venda)
class VendaAdmin(HistoryAdminBase):
    list_display = ('data_venda', 'cliente', 'vendedor', 'calcular_valor_total', 'status_contrato', 'is_deleted', 'is_trocado', 'loja')
    list_filter = (
        'loja',
        'status_contrato',
        'is_deleted',
        'is_trocado',
        'data_venda',
        'vendedor',
        'caixa',
    )
    search_fields = (
        'id',
        'cliente__nome',
        'cliente__cpf',
        'cliente__telefone',
        'cliente__telefone_secundario',
        'vendedor__first_name',
        'vendedor__last_name',
        'vendedor__username',
        'loja__nome',
    )
    date_hierarchy = 'data_venda'
    ordering = ('-data_venda',)
    autocomplete_fields = ('cliente', 'vendedor', 'caixa', 'loja')
    list_select_related = ('cliente', 'vendedor', 'loja', 'caixa')
    list_per_page = 50
    inlines = [ProdutoVendaInline, PagamentoInline]
    readonly_fields = AdminBase.readonly_fields + ('contrato_publico_uuid', 'contrato_publico_link', 'data_venda')
    fieldsets = (
        (None, {
            'fields': (
                'cliente',
                'vendedor',
                'caixa',
                'observacao',
                'repasse_logista',
                'documento_assinado',
                'foto_cliente',
                'imagem_imei',
                'loja',
            )
        }),
        ('Status', {
            'fields': (
                'is_deleted',
                'is_trocado',
            )
        }),
        ('Contrato Publico', {
            'fields': (
                'status_contrato',
                'contrato_publico_uuid',
                'contrato_publico_link',
            )
        }),
        ('Auditoria', {
            'fields': ('data_venda', 'criado_em', 'modificado_em')
        }),
    )

    def contrato_publico_link(self, obj):
        if not obj.pk or not obj.contrato_publico_uuid:
            return 'Link ainda nao gerado.'
        url = reverse("api-contrato-publico", kwargs={"token": obj.contrato_publico_uuid})
        return format_html('<a href="{}" target="_blank" rel="noopener noreferrer">{}</a>', url, url)
    contrato_publico_link.short_description = 'Link do contrato publico'

@admin.register(Pagamento)
class PagamentoAdmin(HistoryAdminBase):
    list_display = (
        'venda',
        'tipo_pagamento',
        'valor',
        'parcelas',
        'valor_parcela',
        'data_primeira_parcela',
        'quitado',
        'bloqueado',
        'flag_atrasado',
        'devolucao',
        'loja',
    )
    list_filter = (
        'tipo_pagamento',
        'loja',
        'quitado',
        'bloqueado',
        'desativado',
        'devolucao',
        'flag_atrasado',
        'sem_contato',
        'mais_prazo',
        'bo',
        'sem_conexao',
        'roubo',
        'lembrete',
        'statuses',
        'data_primeira_parcela',
    )
    search_fields = (
        'id',
        'venda__id',
        'venda__cliente__nome',
        'venda__cliente__cpf',
        'venda__vendedor__first_name',
        'venda__vendedor__last_name',
        'venda__vendedor__username',
    )
    date_hierarchy = 'data_primeira_parcela'
    ordering = ('-data_primeira_parcela',)
    autocomplete_fields = ('venda', 'tipo_pagamento', 'loja')
    list_select_related = ('venda', 'venda__cliente', 'tipo_pagamento', 'loja')
    list_per_page = 50
    filter_horizontal = ('statuses',)

@admin.register(TipoPagamento)
class TipoPagamentoAdmin(AdminBase):
    list_display = ('nome', 'caixa', 'parcelas', 'financeira', 'carne', 'nao_contabilizar', 'loja')
    list_filter = ('caixa', 'parcelas', 'financeira', 'carne', 'nao_contabilizar', 'loja')
    search_fields = ('nome',)
    ordering = ('nome',)
    
@admin.register(Caixa)
class CaixaAdmin(AdminBase):
    list_display = ('data_abertura', 'data_fechamento', 'loja', 'caixa_fechado', 'quantidade_vendas')
    list_filter = ('loja', 'data_abertura', 'data_fechamento')
    search_fields = ('loja__nome', 'id')
    date_hierarchy = 'data_abertura'
    ordering = ('-data_abertura',)
    autocomplete_fields = ('loja',)
    list_select_related = ('loja',)
    
@admin.register(ProdutoVenda)
class ProdutoVendaAdmin(admin.ModelAdmin):
    list_display = ('produto', 'quantidade', 'venda', 'loja')
    list_filter = ('loja',)
    search_fields = ('produto__nome',)


class VendaClienteInline(admin.TabularInline):
    model = Venda
    fk_name = 'cliente'
    extra = 0
    can_delete = False
    show_change_link = True
    verbose_name = 'Venda vinculada'
    verbose_name_plural = 'Vendas vinculadas'
    fields = ('data_venda', 'vendedor', 'status_contrato', 'is_deleted', 'is_trocado', 'repasse_logista', 'loja')
    readonly_fields = fields
    ordering = ('-data_venda',)

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('vendedor', 'loja')


@admin.register(Cliente)
class ClienteAdmin(HistoryAdminBase):
    list_display = ('nome', 'cpf', 'telefone', 'telefone_secundario', 'email', 'cidade', 'profissao', 'recebe_auxilio', 'total_renda', 'loja', 'criado_em')
    list_filter = (
        'loja',
        'recebe_auxilio',
        'cidade',
        'criado_em',
    )
    search_fields = (
        'nome',
        'cpf',
        'rg',
        'telefone',
        'telefone_secundario',
        'email',
        'cidade',
        'bairro',
        'cep',
    )
    date_hierarchy = 'criado_em'
    ordering = ('-criado_em',)
    autocomplete_fields = ('loja',)
    raw_id_fields = ('comprovantes', 'contato_adicional', 'informacao_pessoal')
    list_select_related = ('loja',)
    list_per_page = 50
    inlines = [VendaClienteInline]

@admin.register(Endereco)
class EnderecoAdmin(AdminBase):
    list_display = ('numero', 'bairro', 'cidade', 'cep')

@admin.register(ComprovantesCliente)
class ComprovantesClienteAdmin(HistoryAdminBase):
    pass


@admin.register(ConsultaSerasaAcesso)
class ConsultaSerasaAcessoAdmin(admin.ModelAdmin):
    list_display = (
        'aberto_em',
        'cliente_nome',
        'cliente_loja',
        'tipo',
        'usuario_display',
        'ip_address',
    )
    list_filter = ('tipo', 'aberto_em', 'cliente__loja')
    search_fields = (
        'cliente__nome',
        'cliente__cpf',
        'usuario__username',
        'usuario__first_name',
        'usuario__last_name',
        'ip_address',
    )
    readonly_fields = ('cliente', 'tipo', 'usuario', 'aberto_em', 'ip_address', 'user_agent')
    date_hierarchy = 'aberto_em'
    list_select_related = ('cliente', 'cliente__loja', 'usuario')
    actions = ('exportar_excel', 'exportar_pdf')

    def has_add_permission(self, request):
        return False

    def cliente_nome(self, obj):
        return obj.cliente.nome if obj.cliente else '—'
    cliente_nome.short_description = 'Cliente'
    cliente_nome.admin_order_field = 'cliente__nome'

    def cliente_loja(self, obj):
        return obj.cliente.loja.nome if obj.cliente and obj.cliente.loja else '—'
    cliente_loja.short_description = 'Loja'
    cliente_loja.admin_order_field = 'cliente__loja__nome'

    def usuario_display(self, obj):
        if not obj.usuario:
            return '—'
        nome = (obj.usuario.first_name or '').strip()
        sobrenome = (obj.usuario.last_name or '').strip()
        full = f"{nome} {sobrenome}".strip()
        return full or obj.usuario.username
    usuario_display.short_description = 'Usuário'
    usuario_display.admin_order_field = 'usuario__username'

    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(request, extra_context=extra_context)
        try:
            qs = response.context_data['cl'].queryset
        except (AttributeError, KeyError):
            return response
        extra = {
            'total_aberturas': qs.count(),
            'total_serasa': qs.filter(tipo='serasa').count(),
            'total_serasa_2': qs.filter(tipo='serasa_2').count(),
            'usuarios_unicos': qs.exclude(usuario__isnull=True).values('usuario').distinct().count(),
        }
        response.context_data.update(extra)
        return response

    @admin.action(description='📊 Exportar selecionados para Excel')
    def exportar_excel(self, request, queryset):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.message_user(request, 'openpyxl não está instalado.', level=messages.ERROR)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = 'Acessos Serasa'

        headers = ['Data/Hora', 'Cliente', 'CPF', 'Loja', 'Tipo', 'Usuário', 'Username', 'IP', 'User-Agent']
        ws.append(headers)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1E3A8A')
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        qs = queryset.select_related('cliente', 'cliente__loja', 'usuario')
        for a in qs:
            usuario_nome = '—'
            username = '—'
            if a.usuario:
                nome = (a.usuario.first_name or '').strip()
                sobrenome = (a.usuario.last_name or '').strip()
                usuario_nome = f"{nome} {sobrenome}".strip() or a.usuario.username
                username = a.usuario.username
            ws.append([
                timezone.localtime(a.aberto_em).strftime('%d/%m/%Y %H:%M:%S'),
                a.cliente.nome if a.cliente else '—',
                getattr(a.cliente, 'cpf', '—') or '—',
                a.cliente.loja.nome if a.cliente and a.cliente.loja else '—',
                a.get_tipo_display(),
                usuario_nome,
                username,
                a.ip_address or '—',
                (a.user_agent or '')[:200],
            ])

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        filename = f"acessos_serasa_{timezone.now():%Y%m%d_%H%M%S}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @admin.action(description='📄 Exportar selecionados para PDF')
    def exportar_pdf(self, request, queryset):
        try:
            from weasyprint import HTML
        except ImportError:
            self.message_user(request, 'weasyprint não está instalado.', level=messages.ERROR)
            return

        qs = queryset.select_related('cliente', 'cliente__loja', 'usuario').order_by('-aberto_em')
        linhas = []
        for a in qs:
            usuario_nome = '—'
            if a.usuario:
                nome = (a.usuario.first_name or '').strip()
                sobrenome = (a.usuario.last_name or '').strip()
                usuario_nome = f"{nome} {sobrenome}".strip() or a.usuario.username
            linhas.append({
                'data': timezone.localtime(a.aberto_em).strftime('%d/%m/%Y %H:%M'),
                'cliente': a.cliente.nome if a.cliente else '—',
                'cpf': getattr(a.cliente, 'cpf', '—') or '—',
                'loja': a.cliente.loja.nome if a.cliente and a.cliente.loja else '—',
                'tipo': a.get_tipo_display(),
                'usuario': usuario_nome,
                'ip': a.ip_address or '—',
            })

        html_string = render_to_string('vendas/admin/relatorio_serasa_pdf.html', {
            'linhas': linhas,
            'total': len(linhas),
            'total_serasa': sum(1 for l in linhas if 'Consulta Serasa 2' not in l['tipo'] and 'Serasa' in l['tipo']),
            'total_serasa_2': sum(1 for l in linhas if 'Consulta Serasa 2' in l['tipo']),
            'gerado_em': timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M'),
        })
        pdf_bytes = HTML(string=html_string).write_pdf()
        filename = f"acessos_serasa_{timezone.now():%Y%m%d_%H%M%S}.pdf"
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@admin.register(Loja)
class LojaAdmin(AdminBase):
    list_display = ('nome', 'cnpj', 'telefone', 'credfacil', 'pode_vender_iphone')
    list_filter = ('credfacil', 'pode_vender_iphone')
    search_fields = ('nome', 'cnpj', 'telefone', 'inscricao_estadual')
    ordering = ('nome',)
    list_per_page = 50

    def get_readonly_fields(self, request, obj=None):
        readonly_fields = super().get_readonly_fields(request, obj)
        # Apenas ADMINISTRADOR pode editar pode_vender_iphone
        if not request.user.is_superuser and not request.user.groups.filter(name='ADMINISTRADOR').exists():
            readonly_fields = list(readonly_fields) + ['pode_vender_iphone']
        return readonly_fields
    
    
@admin.register(Parcela)
class ParcelaAdmin(HistoryAdminBase):
    list_display = (
        'pagamento',
        'numero_parcela',
        'valor',
        'valor_pago',
        'desconto',
        'data_vencimento',
        'data_pagamento',
        'pago',
        'tipo_pagamento',
        'loja',
    )
    list_filter = (
        'pago',
        'pagamento_efetuado',
        'tipo_pagamento',
        'loja',
        'data_vencimento',
        'data_pagamento',
    )
    search_fields = (
        'pagamento__id',
        'pagamento__venda__id',
        'pagamento__venda__cliente__nome',
        'pagamento__venda__cliente__cpf',
    )
    date_hierarchy = 'data_vencimento'
    ordering = ('-data_vencimento',)
    autocomplete_fields = ('pagamento', 'tipo_pagamento', 'loja')
    list_select_related = ('pagamento', 'pagamento__venda', 'tipo_pagamento', 'loja')
    list_per_page = 50


@admin.register(LancamentoCaixa)
class LancamentoCaixaAdmin(AdminBase):
    list_display = ('caixa', 'tipo_lancamento', 'valor')
    
    
@admin.register(AnaliseCreditoCliente)
class AnaliseCreditoClienteAdmin(HistoryAdminBase):
    list_display = (
        'cliente',
        'produto',
        'status',
        'status_aplicativo',
        'analise_online',
        'data_analise',
        'data_aprovacao',
        'aprovado_por',
        'venda',
        'loja',
    )
    list_filter = (
        'status',
        'status_aplicativo',
        'analise_online',
        'loja',
        'produto',
        'aprovado_por',
        'icloud_configurado_vendedor',
        'icloud_confirmado_analista',
        'data_analise',
        'data_aprovacao',
    )
    search_fields = (
        'cliente__nome',
        'cliente__cpf',
        'cliente__telefone',
        'cliente__telefone_secundario',
        'produto__nome',
        'imei__imei',
        'imei_informado',
        'loja__nome',
    )
    date_hierarchy = 'data_analise'
    ordering = ('-data_analise',)
    autocomplete_fields = ('cliente', 'produto', 'imei', 'venda', 'aprovado_por', 'numero_autenticador', 'loja')
    list_select_related = ('cliente', 'produto', 'aprovado_por', 'venda', 'loja')
    list_per_page = 50
    list_editable = ('status',)


@admin.register(StatusPagamento)
class StatusPagamentoAdmin(AdminBase):
    list_display = ('nome', 'slug', 'cor_hex')
    search_fields = ('nome',)
    list_editable = ('cor_hex',)


class NumeroAutenticadorForm(forms.ModelForm):
    numero = forms.CharField(
        max_length=20,
        widget=forms.TextInput(attrs={
            'class': 'vTextField',
            'placeholder': '(00) 00000-0000',
            'data-mask': '(00) 00000-0000',
        }),
        label='Número de Telefone',
    )

    class Meta:
        model = NumeroAutenticador
        fields = '__all__'


@admin.register(NumeroAutenticador)
class NumeroAutenticadorAdmin(AdminBase):
    form = NumeroAutenticadorForm
    list_display = ('numero', 'descricao', 'ativo', 'loja')
    list_filter = ('ativo', 'loja')
    search_fields = ('numero', 'descricao')
    list_editable = ('ativo',)

    class Media:
        js = ('admin/js/numero_autenticador_mask.js',)


@admin.register(PreAnaliseRapida)
class PreAnaliseRapidaAdmin(AdminBase):
    list_display = ('nome_completo', 'cpf', 'status', 'tem_comprovante_residencia',
                    'possui_duas_referencias', 'loja', 'criado_por', 'analisado_por', 'criado_em')
    list_filter = ('status', 'tem_comprovante_residencia', 'possui_duas_referencias', 'loja', 'criado_em')
    search_fields = ('nome_completo', 'cpf', 'loja__nome')
    date_hierarchy = 'criado_em'
    ordering = ('-criado_em',)
    readonly_fields = ('criado_em', 'modificado_em', 'data_decisao', 'analisado_por', 'cliente_gerado')
