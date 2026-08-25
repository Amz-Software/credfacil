from io import BytesIO
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP

import weasyprint
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.db import transaction
from django.http import HttpResponse
from django.db.models import OuterRef, Subquery, DateField, Q, Sum, Max, Prefetch
from django.shortcuts import get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.generic import CreateView, UpdateView, ListView, DetailView, TemplateView, View

from accounts.views import logout_view
from financeiro.forms import *
from financeiro.forms import RelatorioSaidaForm
from vendas.forms import ContatoForm
from vendas.models import Contato, Loja, Parcela, Pagamento, StatusPagamento
from vendas.views import BaseView
from .models import CaixaMensal, CaixaMensalFuncionario, CaixaMensalGastoFixo, GastoFixo, GastosAleatorios


def _mask_cpf(cpf):
    digits = ''.join(ch for ch in (cpf or '') if ch.isdigit())
    if len(digits) != 11:
        return cpf or '—'
    return f'{digits[:3]}.***.***-{digits[-2:]}'


def _format_date(value):
    if not value:
        return None
    if hasattr(value, 'strftime'):
        return value.strftime('%d/%m/%Y')
    parsed = parse_date(str(value))
    return parsed.strftime('%d/%m/%Y') if parsed else value


def _build_references_text(cliente):
    references = []

    for contato in cliente.contatos.all().order_by('data'):
        if contato.observacao:
            references.append(contato.observacao.strip())

    if cliente.informacao_pessoal:
        info = cliente.informacao_pessoal
        partes = [p for p in [info.nome, info.contato, info.endereco] if p]
        if partes:
            references.append(' - '.join([partes[0], ', '.join(partes[1:])]).strip(' -'))

    if cliente.contato_adicional:
        info = cliente.contato_adicional
        partes = [p for p in [info.nome_adicional, info.contato, info.endereco_adicional] if p]
        if partes:
            references.append(' - '.join([partes[0], ', '.join(partes[1:])]).strip(' -'))

    if not references:
        references = [
            'I. [Nome ou inicial], [vínculo], reside no endereço [endereço].',
            'II. [Nome ou inicial], [vínculo], telefone [telefone], observação [texto].',
        ]

    return '\n'.join(f'{idx + 1}. {texto}' for idx, texto in enumerate(references))


def _get_primeira_parcela_atrasada(pagamento):
    hoje = timezone.localdate()
    return (
        pagamento.parcelas_pagamento
        .filter(pago=False, data_vencimento__lt=hoje)
        .order_by('data_vencimento', 'numero_parcela')
        .first()
    )


def _get_notificacao_bo_context(pagamento, user, referencias_personais):
    venda = pagamento.venda
    cliente = venda.cliente
    loja = venda.loja
    item = venda.itens_venda.select_related('produto__marca').first()
    parcelas = list(pagamento.parcelas_pagamento.all().order_by('data_vencimento', 'numero_parcela'))
    primeira_parcela_atrasada = _get_primeira_parcela_atrasada(pagamento)
    contrato_meta = venda.loja.contrato or {}

    inicio_vigencia = _format_date(
        contrato_meta.get('data_inicio') or contrato_meta.get('inicio_vigencia') or venda.data_venda.date()
    )
    fim_vigencia = _format_date(
        contrato_meta.get('data_fim')
        or contrato_meta.get('fim_vigencia')
        or (parcelas[-1].data_vencimento if parcelas else venda.data_venda.date())
    )

    produto = item.produto if item else None
    telefone_loja = getattr(loja, 'telefone', None) or contrato_meta.get('telefone') or '—'
    cidade_loja = getattr(loja, 'cidade', None) or contrato_meta.get('cidade') or '—'
    uf_loja = getattr(loja, 'uf', None) or contrato_meta.get('uf') or '—'
    endereco_loja = getattr(loja, 'endereco', None) or contrato_meta.get('endereco') or '—'
    endereco_cliente = ', '.join(part for part in [cliente.endereco, cliente.bairro, cliente.cidade] if part)
    if cidade_loja != '—' and uf_loja != '—':
        bo_local_assinatura = f'{cidade_loja}/{uf_loja}, {timezone.localtime().strftime("%d/%m/%Y")}'
    else:
        bo_local_assinatura = f'{loja.nome}, {timezone.localtime().strftime("%d/%m/%Y")}'

    return {
        'pagamento': pagamento,
        'venda': venda,
        'cliente': cliente,
        'loja': loja,
        'item': item,
        'produto': produto,
        'primeira_parcela_atrasada': primeira_parcela_atrasada,
        'referencias_personais': referencias_personais,
        'cliente_cpf_exibicao': _mask_cpf(cliente.cpf),
        'cliente_endereco': endereco_cliente or '—',
        'loja_nome': loja.nome,
        'loja_cnpj': loja.cnpj or '—',
        'loja_telefone': telefone_loja,
        'loja_endereco': endereco_loja,
        'loja_cidade': cidade_loja,
        'loja_uf': uf_loja,
        'loja_logo': loja.logo_loja,
        'data_assinatura': venda.data_venda.date(),
        'inicio_vigencia': inicio_vigencia,
        'fim_vigencia': fim_vigencia,
        'produto_nome': produto.nome if produto else '—',
        'produto_modelo': produto.nome if produto else '—',
        'produto_imei': item.imei if item and item.imei else '—',
        'valor_mensal': primeira_parcela_atrasada.valor if primeira_parcela_atrasada else pagamento.valor_parcela,
        'vencimento_primeira_atrasada': primeira_parcela_atrasada.data_vencimento if primeira_parcela_atrasada else None,
        'created_at': timezone.localtime(),
        'bo_local_assinatura': bo_local_assinatura,
    }


def _get_notificacao_bo_session_key(pagamento_id):
    return f'notificacao_bo_referencias_{pagamento_id}'


def _render_notificacao_bo_pdf(request, pagamento, referencias_personais):
    context = _get_notificacao_bo_context(pagamento, request.user, referencias_personais)
    html = render_to_string('notificacao_bo/notificacao_bo_pdf.html', context, request=request)
    pdf_buffer = BytesIO()
    weasyprint.HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf(pdf_buffer)
    response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="notificacao_bo_pagamento_{pagamento.pk}.pdf"'
    return response


class NotificacaoBoPagamentoMixin(PermissionRequiredMixin):
    permission_required = 'financeiro.add_notificacaobo'

    def dispatch(self, request, *args, **kwargs):
        self.pagamento = get_object_or_404(
            Pagamento.objects.select_related('venda__cliente', 'venda__loja')
            .prefetch_related('parcelas_pagamento', 'venda__itens_venda__produto__marca', 'venda__cliente__contatos'),
            pk=kwargs.get('pk')
        )
        if not _get_primeira_parcela_atrasada(self.pagamento):
            messages.error(request, 'Não há parcela vencida em aberto para gerar o BO.')
            return redirect('financeiro:contas_a_receber_update', pk=self.pagamento.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_referencias_padrao(self):
        return _build_references_text(self.pagamento.venda.cliente)


class GerarNotificacaoBoView(NotificacaoBoPagamentoMixin, TemplateView):
    template_name = 'notificacao_bo/notificacao_bo_preview.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = kwargs.get('form') or NotificacaoBoForm(initial={'referencias_personais': self.get_referencias_padrao()})
        referencias_personais = form['referencias_personais'].value()
        context.update(_get_notificacao_bo_context(self.pagamento, self.request.user, referencias_personais))
        context['form'] = form
        return context

    def post(self, request, *args, **kwargs):
        form = NotificacaoBoForm(request.POST)
        if not form.is_valid():
            return self.render_to_response(self.get_context_data(form=form))

        referencias_personais = form.cleaned_data.get('referencias_personais', '').strip() or self.get_referencias_padrao()
        request.session[_get_notificacao_bo_session_key(self.pagamento.pk)] = referencias_personais
        return redirect('financeiro:notificacao_bo_pdf', pk=self.pagamento.pk)


class GerarNotificacaoBoPdfView(NotificacaoBoPagamentoMixin, View):
    def get(self, request, *args, **kwargs):
        referencias_personais = request.session.get(
            _get_notificacao_bo_session_key(self.pagamento.pk),
            self.get_referencias_padrao(),
        )
        return _render_notificacao_bo_pdf(request, self.pagamento, referencias_personais)


class CaixaMensalListView(BaseView, PermissionRequiredMixin, ListView):
    model = CaixaMensal
    template_name = 'caixa_mensal/caixa_mensal_list.html'
    context_object_name = 'caixas_mensais'
    paginate_by = 10
    permission_required = 'financeiro.view_caixamensal'
    def get_queryset(self):
        loja_id = self.request.session.get('loja_id')
        loja = Loja.objects.get(id=loja_id)
        queryset = super().get_queryset()
        search = self.request.GET.get('search', '')
        if search:
            data = search + "-01"
            data = datetime.strptime(data, "%Y-%m-%d").date()
            queryset = queryset.filter(mes__month=data.month, mes__year=data.year, loja=loja).order_by('-mes')
        return queryset.filter(loja=loja).order_by('-mes')


@login_required
@permission_required('financeiro.add_caixamensal', raise_exception=True)
def caixa_mensal_create(request):
    with transaction.atomic():
        
        loja_id = request.session.get('loja_id')
        mes = request.POST.get('mes')

        if mes:
            mes = mes + "-01"
            mes_atual = datetime.strptime(mes, "%Y-%m-%d").date()
        
        if not loja_id:
            messages.error(request, "Loja não selecionada. Selecione uma loja e faça login novamente.")
            return logout_view(request)

        try:
            loja = Loja.objects.get(id=loja_id)
        except Loja.DoesNotExist:
            messages.error(request, "Loja não encontrada. Selecione uma loja e faça login novamente.")
            return logout_view(request)

        if CaixaMensal.objects.filter(loja=loja, mes=mes_atual).exists():
            messages.error(request, "Caixa mensal já criado para este mês.")
            return redirect('financeiro:caixa_mensal_list')

        # Cria o caixa mensal
        caixa_mensal = CaixaMensal.objects.create(
            loja=loja,
            mes=mes_atual,
            valor=0.00,
        )

        # Associar todos os funcionários e gastos fixos
        funcionarios = loja.usuarios.all()
        for funcionario in funcionarios:
            CaixaMensalFuncionario.objects.create(
                caixa_mensal=caixa_mensal,
                funcionario=funcionario,
                salario=0.00,  # Inicializa com valores zero para edição posterior
                comissao=0.00,
            )
        
        gastos_fixos = GastoFixo.objects.filter(loja=loja)
        if gastos_fixos:
            for gasto in gastos_fixos:
                CaixaMensalGastoFixo.objects.create(
                    caixa_mensal=caixa_mensal,
                    gasto_fixo=gasto,
                    valor=0.00,  # Inicializa com valores zero para edição posterior
                    observacao="",
                )
        
        # Redirecionar para a página de detalhes do caixa mensal recém-criado
        messages.success(request, "Caixa mensal criado com sucesso.")
        return redirect('financeiro:caixa_mensal_update', pk=caixa_mensal.pk)

@login_required
@permission_required('financeiro.change_caixamensal', raise_exception=True)
def fechar_caixa_mensal(request, pk):
    caixa_mensal = get_object_or_404(CaixaMensal, pk=pk)
    caixa_mensal.fechar()
    return redirect('financeiro:caixa_mensal_list')

@login_required
@permission_required('financeiro.change_caixamensal', raise_exception=True)
def reabrir_caixa_mensal(request, pk):
    caixa_mensal = get_object_or_404(CaixaMensal, pk=pk)
    caixa_mensal.reabrir()
    return redirect('financeiro:caixa_mensal_list')



# class CaixaMensalCreateView(CreateView):
#     model = CaixaMensal
#     template_name = 'caixa_mensal/caixa_mensal_form.html'
#     fields = ['valor']  # Apenas o campo inicial, loja e mês são automáticos

#     def form_valid(self, form): 
#         form.instance.mes = datetime.now()  # Define o mês atual
#         return super().form_valid(form)


# class CaixaMensalDetailView(DetailView):
#     model = CaixaMensal
#     template_name = 'caixa_mensal/caixa_mensal_form.html'

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         self.object = self.get_object()
#         context['object'] = self.object

#         # Usando Prefetch para carregar os gastos fixos e funcionários relacionados à caixa mensal
#         gastos_fixos_qs = GastoFixo.objects.filter(
#             lojas_gastos_fixos__caixa_mensal=self.object
#         )
#         funcionarios_qs = Funcionario.objects.filter(
#             lojas_funcionarios__caixa_mensal=self.object
#         )

#         # Definindo as consultas para as tabelas intermediárias
#         context['gastos_fixos'] = gastos_fixos_qs
#         context['funcionarios'] = funcionarios_qs

#         # Configurando o formset de Gastos Aleatórios
#         GastosAleatoriosFormSet = inlineformset_factory(
#             CaixaMensal, GastosAleatorios,
#             form=GastosAleatoriosForm,
#             extra=0, can_delete=False
#         )

#         if self.request.method == 'POST':
#             context['gastos_aleatorios_formset'] = GastosAleatoriosFormSet(self.request.POST, instance=self.object)
#         else:
#             context['gastos_aleatorios_formset'] = GastosAleatoriosFormSet(instance=self.object)

#         return context

#     def post(self, request, *args, **kwargs):
#         self.object = self.get_object()
#         context = self.get_context_data()

#         # Processando o formulário de Gastos Fixos
#         if 'gastos_fixos_form' in request.POST:
#             for gasto_fixo in self.object.gastos_fixos.all():
#                 valor = request.POST.get(f'valor_{gasto_fixo.id}')
#                 observacao = request.POST.get(f'observacao_{gasto_fixo.id}')
#                 if valor is not None:
#                     gasto_fixo.valor = float(valor) if valor else 0
#                 if observacao is not None:
#                     gasto_fixo.observacao = observacao
#                 gasto_fixo.save()
#             messages.success(request, 'Gastos Fixos atualizados com sucesso.')

#         # Processando o formulário de Funcionários
#         if 'funcionario_form' in request.POST:
#             for funcionario in self.object.funcionarios.all():
#                 salario = request.POST.get(f'salario_{funcionario.id}')
#                 comissao = request.POST.get(f'comissao_{funcionario.id}')
#                 if salario is not None:
#                     funcionario.salario = float(salario) if salario else 0
#                 if comissao is not None:
#                     funcionario.comissao = float(comissao) if comissao else 0
#                 funcionario.save()
#             messages.success(request, 'Funcionários atualizados com sucesso.')

#         # Processando o formset de Gastos Aleatórios
#         gastos_aleatorios_formset = context['gastos_aleatorios_formset']
#         if gastos_aleatorios_formset.is_valid():
#             gastos_aleatorios_formset.save()
#             messages.success(request, 'Gastos Aleatórios atualizados com sucesso.')
#         else:
#             messages.error(request, 'Erro ao atualizar os Gastos Aleatórios.')

#         return HttpResponseRedirect(self.object.get_absolute_url())

class CaixaMensalDetailView(PermissionRequiredMixin, DetailView):
    model = CaixaMensal
    template_name = 'caixa_mensal/caixa_mensal_detail.html'
    context_object_name = 'caixa_mensal'
    permission_required = 'financeiro.view_caixamensal'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loja_id = self.request.session.get('loja_id')
        caixa_mensal = self.get_object()
        caixa_mensal_gastos_fixos = CaixaMensalGastoFixo.objects.filter(caixa_mensal=caixa_mensal)
        caixa_mensal_funcionarios = CaixaMensalFuncionario.objects.filter(caixa_mensal=caixa_mensal)
        caixa_mensal_gastos_aleatorios = GastosAleatorios.objects.filter(caixa_mensal=caixa_mensal)

        # Pré-configurar os Gastos Fixos associados ao Caixa Mensal
        self._pre_configurar_gastos_fixos(caixa_mensal)

        # Pré-configurar os Funcionários associados ao Caixa Mensal
        self._pre_configurar_funcionarios(caixa_mensal)

        # Formsets com prefixos
        context['formset_gastos_fixos'] = CaixaMensalGastoFixoFormSet(
            queryset=CaixaMensalGastoFixo.objects.filter(caixa_mensal=caixa_mensal),
            prefix='gastos_fixos'
        )
        context['formset_funcionarios'] = CaixaMensalFuncionarioFormSet(
            queryset=CaixaMensalFuncionario.objects.filter(caixa_mensal=caixa_mensal),
            prefix='funcionarios'
        )
        context['formset_gastos_aleatorios'] = GastosAleatoriosFormSet(
            instance=caixa_mensal,
            prefix='gastos_aleatorios'
        )

        caixas_mes = Caixa.objects.filter(loja_id=loja_id).filter(data_abertura__month=caixa_mensal.mes.month)
        vendas = []

        total_saidas = 0

        for caixa in caixas_mes:
            vendas += caixa.vendas.all()
            total_saidas += caixa.saidas

        total_custo = 0
        total_lucro = 0
        valor_venda_por_tipo_pagamento = {}

        for venda in vendas:
            total_custo += venda.custo_total()
            total_lucro += venda.lucro_total()
            #filtrar pagamentos para o mes atual
            for pagamento in venda.pagamentos.all():
                if not pagamento.tipo_pagamento.nao_contabilizar:
                    if pagamento.tipo_pagamento.nome not in valor_venda_por_tipo_pagamento:
                        valor_venda_por_tipo_pagamento[pagamento.tipo_pagamento.nome] = 0
                    valor_venda_por_tipo_pagamento[pagamento.tipo_pagamento.nome] += pagamento.valor 
        valor_por_tipo_pagamento_total = sum(valor_venda_por_tipo_pagamento.values())

        lucro_total = (valor_por_tipo_pagamento_total - total_custo)

        total_gasto_fixos = sum(gasto.valor for gasto in caixa_mensal_gastos_fixos)
        total_funcionarios = sum(funcionario.salario + funcionario.comissao for funcionario in caixa_mensal_funcionarios)
        total_gastos_aleatorios = sum(gasto.valor for gasto in caixa_mensal_gastos_aleatorios)
        total_gastos = total_gasto_fixos + total_funcionarios + total_gastos_aleatorios

        saldo_final = (caixa_mensal.valor + lucro_total) - total_gastos - total_saidas

        context = {
            'total_vendas': valor_por_tipo_pagamento_total,
            'valor_venda_por_tipo_pagamento': valor_venda_por_tipo_pagamento.items(),
            'valor_por_tipo_pagamento_total': valor_por_tipo_pagamento_total,
            'total_custo': total_custo,
            'total_lucro': total_lucro,
            'lucro_total': lucro_total,
            'total_saidas': total_saidas,
            'total_gasto_fixos': total_gasto_fixos,
            'total_funcionarios': total_funcionarios,
            'total_gastos_aleatorios': total_gastos_aleatorios,
            'saldo_final': saldo_final,
            'formset_gastos_fixos': context['formset_gastos_fixos'],
            'formset_funcionarios': context['formset_funcionarios'],
            'formset_gastos_aleatorios': context['formset_gastos_aleatorios'],
            'caixa_mensal': caixa_mensal
        }

        return context

    def _pre_configurar_gastos_fixos(self, caixa_mensal):
        """Cria associações de Gastos Fixos ao Caixa Mensal, caso não existam."""
        if not CaixaMensalGastoFixo.objects.filter(caixa_mensal=caixa_mensal).exists():
            gastos_fixos_disponiveis = GastoFixo.objects.all().filter(loja=caixa_mensal.loja)
            for gasto_fixo in gastos_fixos_disponiveis:
                CaixaMensalGastoFixo.objects.create(
                    caixa_mensal=caixa_mensal,
                    gasto_fixo=gasto_fixo,
                    valor=0.00,  # Valor padrão
                    observacao=""
                )

    def _pre_configurar_funcionarios(self, caixa_mensal):
        """Cria associações de Funcionários ao Caixa Mensal, caso não existam."""
        if not CaixaMensalFuncionario.objects.filter(caixa_mensal=caixa_mensal).exists():
            funcionarios_disponiveis = Loja.objects.get(id=self.request.session.get('loja_id')).usuarios.all()
            for funcionario in funcionarios_disponiveis:
                CaixaMensalFuncionario.objects.create(
                    caixa_mensal=caixa_mensal,
                    funcionario=funcionario,
                    salario=0.00,  # Valor padrão
                    comissao=0.00
                )

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        caixa_mensal = self.object

        # Inicialização dos formsets com prefixos
        formset_gastos_fixos = CaixaMensalGastoFixoFormSet(
            request.POST,
            queryset=CaixaMensalGastoFixo.objects.filter(caixa_mensal=caixa_mensal),
            prefix='gastos_fixos'
        )
        formset_funcionarios = CaixaMensalFuncionarioFormSet(
            request.POST,
            queryset=CaixaMensalFuncionario.objects.filter(caixa_mensal=caixa_mensal),
            prefix='funcionarios'
        )
        formset_gastos_aleatorios = GastosAleatoriosFormSet(
            request.POST,
            instance=caixa_mensal,
            prefix='gastos_aleatorios'
        )

        # Verificar se os formsets são válidos
        if formset_gastos_fixos.is_valid() and formset_funcionarios.is_valid() and formset_gastos_aleatorios.is_valid():
            # Salvar formset de Gastos Fixos
            instances_gastos_fixos = formset_gastos_fixos.save(commit=False)
            for instance in instances_gastos_fixos:
                instance.caixa_mensal = caixa_mensal
                instance.save()
            formset_gastos_fixos.save_m2m()

            # Salvar formset de Funcionários
            # verificar se o funcionário foi deletado
            for form in formset_funcionarios.deleted_forms:
                if form.instance.pk:
                    form.instance.delete()
            instances_funcionarios = formset_funcionarios.save(commit=False)
            
            for instance in instances_funcionarios:
                instance.caixa_mensal = caixa_mensal
                instance.save(user=request.user)
            formset_funcionarios.save_m2m()

            # Salvar formset de Gastos Aleatórios
            formset_gastos_aleatorios.save()

            # Mensagem de sucesso e redirecionamento
            messages.success(request, "Caixa Mensal atualizado com sucesso!")
            return redirect('financeiro:caixa_mensal_update', pk=caixa_mensal.pk)

        # Caso haja erros, exibir mensagem de erro e retornar o contexto com os formsets
        messages.error(request, "Erro ao atualizar o Caixa Mensal.")
        print(formset_gastos_fixos.errors)
        print(formset_funcionarios.errors)
        print(formset_gastos_aleatorios.errors)
        return self.render_to_response(self.get_context_data(
            formset_gastos_fixos=formset_gastos_fixos,
            formset_funcionarios=formset_funcionarios,
            formset_gastos_aleatorios=formset_gastos_aleatorios
        ))

class ContasAReceberListView(BaseView, PermissionRequiredMixin, ListView):
    model = Pagamento
    template_name = 'contas_a_receber/contas_a_receber_list.html'
    context_object_name = 'contas_a_receber'
    paginate_by = 10
    permission_required = 'vendas.view_pagamento'

    def get_queryset(self):
        loja_id = self.request.session.get('loja_id')
        user = self.request.user
        qs = Pagamento.objects.exclude(venda__is_deleted=True).with_status_flags()
        
        search = self.request.GET.get('search', '')
        if search:
            qs = qs.filter(
                Q(venda__cliente__nome__icontains=search) |
                Q(venda__cliente__cpf__icontains=search) |
                Q(venda__cliente__contatos__observacao__icontains=search) |
                Q(id__iexact=search) 
            ).distinct()

        if self.request.GET.get('bloqueado'):
            qs = qs.filter(bloqueado=True)

        if self.request.GET.get('desativado'):
            qs = qs.filter(desativado=True)

        if self.request.GET.get('cobranca'):
            qs = qs.filter(flag_atrasado=True)
        
        if self.request.GET.get('pagamento_efetuado'):
            qs = qs.filter(parcelas_pagamento__pagamento_efetuado=True)
        if self.request.GET.get('sem_contato'):
            qs = qs.filter(sem_contato=True)
        if self.request.GET.get('mais_prazo'):
            qs = qs.filter(mais_prazo=True)

        # Filtro por StatusPagamento (M2M)
        status_pagamento_ids = self.request.GET.getlist('status_pagamento')
        if status_pagamento_ids:
            qs = qs.filter(statuses__in=status_pagamento_ids).distinct()

        if not user.has_perm('vendas.can_view_all_payments'):
            qs = qs.filter(loja_id=loja_id)

        status = self.request.GET.get('status')
        if status == 'no_prazo':
            # Filtra clientes que já realizaram pelo menos 1 pagamento
            qs = qs.filter(parcelas_pagamento__pago=True).distinct()
        elif status == 'atrasado':
            qs = qs.filter(com_parcela_atrasada=True)
        elif status == 'quitado':
            qs = qs.filter(todas_parcelas_pagas=True)
        elif status == 'pendente':  # novo
            qs = qs.filter(com_pagamento_pendente=True)

        loja = self.request.GET.get('loja')

        if loja:
            try:
                loja = Loja.objects.get(id=loja)
                qs = qs.filter(loja=loja)
            except Loja.DoesNotExist:
                messages.error(self.request, "Loja não encontrada.")
                return redirect('financeiro:contas_a_receber_list')

            
        return qs.order_by('-criado_em').exclude(tipo_pagamento__caixa=True).filter(tipo_pagamento__parcelas=True)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Usar apenas os itens paginados para evitar quebrar a paginação
        page_items = context.get('contas_a_receber', [])
        for pagamento in page_items:
            status = self.verificar_atraso_parcela(pagamento)
            pagamento.atrasado = status
        if self.request.user.has_perm('vendas.can_view_all_stores'):
            context['lojas'] = Loja.objects.all()
        context['status_pagamento_all'] = StatusPagamento.objects.all()
        context['selected_status_pagamento'] = self.request.GET.getlist('status_pagamento')
        return context

    def verificar_atraso_parcela(self, pagamento):
        data_atual = timezone.now().date()
        parcela = pagamento.parcelas_pagamento.filter(pago=False).order_by('data_vencimento').first()
        if parcela and parcela.data_vencimento < data_atual:
            if parcela.valor_restante < parcela.valor:
                return "Pago parcialmente"
            return "Atrasado"
        return "Em dia"
    


class ContasAReceberDetailView(PermissionRequiredMixin, DetailView):
    model = Pagamento
    template_name = 'contas_a_receber/contas_a_receber_detail.html'
    context_object_name = 'conta_a_receber'
    permission_required = 'vendas.view_pagamento'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        conta_a_receber = self.get_object()
        context['parcela_form'] = ParcelaInlineFormSet(
            instance=conta_a_receber,
            form_kwargs={'user': self.request.user}
        )
        totais_parcelas = conta_a_receber.parcelas_pagamento.aggregate(
            total_valor=Sum('valor'),
            total_valor_pago=Sum('valor_pago'),
        )
        context['total_valor_original'] = totais_parcelas['total_valor'] or 0
        context['total_valor_pago'] = totais_parcelas['total_valor_pago'] or 0
        # Total efetivamente pago (usa valor_pago quando preenchido, senão valor)
        context['total_pago_efetivo'] = sum(
            (p.valor_pago_efetivo for p in conta_a_receber.parcelas_pagamento.filter(pago=True)),
            Decimal('0'),
        )
        # Form Select2 para editar statuses
        context['status_form'] = PagamentoStatusForm(instance=conta_a_receber)
        cliente = self.get_object().venda.cliente
        contatos = list(cliente.contatos.all())
        for c in contatos:
            c.form = ContatoForm(instance=c)
        context['contacts'] = contatos
        context['contato_form'] = ContatoForm()
        context['all_status_pagamento'] = StatusPagamento.objects.all()
        return context


    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        user = request.user

        # 1) Atualiza o desconto, se veio no POST
        if 'porcentagem_desconto' in request.POST:
            if not user.has_perm('vendas.change_pagamento'):
                messages.error(request, "Você não tem permissão para alterar o desconto.")
            else:
                try:
                    novo = Decimal(request.POST['porcentagem_desconto'])
                    self.object.porcentagem_desconto = novo
                    self.object.save(update_fields=['porcentagem_desconto'])
                    messages.success(request, "% de desconto atualizado com sucesso!")
                except Exception:
                    messages.error(request, "Valor de desconto inválido.")
            return redirect(request.path)

        # Atualização de statuses (select múltiplo)
        if 'statuses' in request.POST:
            if not user.has_perm('vendas.change_pagamento'):
                messages.error(request, "Você não tem permissão para alterar status.")
            else:
                form = PagamentoStatusForm(request.POST, instance=self.object)
                if form.is_valid():
                    form.save()
                    messages.success(request, "Status de pagamento atualizados!")
                else:
                    messages.error(request, "Não foi possível atualizar os status.")
            return redirect(request.path)

        # 2) Senão, processa o formset de parcelas
        parcela_form = ParcelaInlineFormSet(
            request.POST,
            instance=self.object,
            form_kwargs={'user': user}
        )

        # Conta parcelas extras (sem pk e não marcadas como DELETE)
        extras_forms = []
        for form in parcela_form.forms:
            if not form.is_valid():
                continue
            if form.cleaned_data.get('DELETE'):
                continue
            if form.instance.pk is None and form.has_changed():
                extras_forms.append(form)

        if extras_forms and not user.has_perm('vendas.add_parcela'):
            messages.error(request, "Você não tem permissão para adicionar parcelas.")
            return redirect(request.path)

        if len(extras_forms) > 4:
            messages.error(request, "É permitido adicionar no máximo 4 parcelas extras por vez.")
            return redirect(request.path)

        if not extras_forms and not user.has_perm('vendas.change_pagamento'):
            messages.error(request, "Você não tem permissão para editar parcelas.")
            return redirect('financeiro:contas_a_receber_list')

        if parcela_form.is_valid():
            # Bloqueia alteração de data de vencimento de parcelas existentes sem permissão
            for form in parcela_form:
                if not form.cleaned_data or form.cleaned_data.get('DELETE'):
                    continue
                if form.instance.pk:
                    original_parcela = Parcela.objects.get(pk=form.instance.pk)
                    if form.cleaned_data.get('data_vencimento') != original_parcela.data_vencimento:
                        if not user.has_perm('vendas.change_vencimento_parcela'):
                            messages.error(request, "Você não tem permissão para alterar a data de vencimento das parcelas.")
                            return redirect(request.path)

            # Quando há parcelas extras, redistribui o saldo (Pagamento.valor − total pago) entre as pendentes
            if extras_forms:
                pago_total = Decimal('0')
                pendentes_forms = []
                for form in parcela_form:
                    if not form.cleaned_data or form.cleaned_data.get('DELETE'):
                        continue
                    if form.cleaned_data.get('pago'):
                        valor_pago = form.cleaned_data.get('valor_pago') or Decimal('0')
                        if not valor_pago:
                            valor_pago = form.cleaned_data.get('valor') or Decimal('0')
                        pago_total += Decimal(valor_pago)
                    else:
                        pendentes_forms.append(form)

                if pendentes_forms:
                    saldo = (Decimal(self.object.valor) - pago_total)
                    if saldo < 0:
                        saldo = Decimal('0')
                    base = (saldo / Decimal(len(pendentes_forms))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    soma_arredondada = base * len(pendentes_forms)
                    diferenca = (saldo - soma_arredondada).quantize(Decimal('0.01'))
                    for idx, form in enumerate(pendentes_forms):
                        valor_aplicado = base + (diferenca if idx == len(pendentes_forms) - 1 else Decimal('0'))
                        form.instance.valor = valor_aplicado
                        form.cleaned_data['valor'] = valor_aplicado

                # Renumera novas parcelas em sequência ao final e garante o FK
                max_numero = self.object.parcelas_pagamento.aggregate(m=Max('numero_parcela'))['m'] or 0
                proxima = max_numero + 1
                for form in extras_forms:
                    form.instance.pagamento = self.object
                    form.instance.numero_parcela = proxima
                    form.cleaned_data['numero_parcela'] = proxima
                    # Reset campos de pagamento para parcelas novas
                    form.instance.pago = False
                    form.instance.valor_pago = 0
                    form.instance.pagamento_efetuado = False
                    form.instance.data_pagamento = None
                    proxima += 1

            for form in parcela_form:
                if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                    if form.instance.pk is None:
                        form.instance.pagamento = self.object
                    form.instance.save(user=user)
            parcela_form.save()
            messages.success(request, "Parcelas atualizadas com sucesso!")
            return redirect(request.path)
        else:
            for form in parcela_form:
                for field, errors in form.errors.items():
                    messages.error(request, f"Erro em {field}: {'; '.join(errors)}")
            messages.error(request, "Erro ao atualizar as parcelas.")
            return self.render_to_response(self.get_context_data(parcela_form=parcela_form))

    
class RelatorioContasAReceberView(BaseView, PermissionRequiredMixin, TemplateView):
    template_name = 'contas_a_receber/relatorio.html'
    permission_required = 'vendas.can_genarate_report_payments'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = RelatorioContasAReceberForm()

        return context

class FolhaRelatorioContasAReceberView(BaseView, PermissionRequiredMixin, TemplateView):
    template_name = 'contas_a_receber/folha.html'
    permission_required = 'vendas.can_genarate_report_payments'
    STATUS_LABELS = dict(RelatorioContasAReceberForm.base_fields['status'].choices)

    def get_context_data(self, **kwargs):
        data_inicio = self.request.GET.get('data_inicial')
        data_fim = self.request.GET.get('data_final')
        lojas = self.request.GET.getlist('lojas')
        status_list = self.request.GET.getlist('status')
        status_pagamento_ids = self.request.GET.getlist('status_pagamento')
        contas_a_receber = []

        if data_inicio and data_fim:
            lojas_qs = Loja.objects.filter(id__in=lojas)
            data_inicio_dt = datetime.strptime(data_inicio, "%Y-%m-%d").date()
            data_final_dt = datetime.strptime(data_fim, "%Y-%m-%d").date()
            data_final_dt_plus = data_final_dt + timedelta(days=1)

            # Caso especial: quando o usuário quer listar apenas os pagos, exiba uma linha por PARCELA paga no período
            # em vez de uma linha por Pagamento. Isso evita ocultar múltiplos pagamentos no mesmo mês.
            if status_list and 'todos' not in status_list and set(status_list) == {'pago'}:
                parcelas_qs = Parcela.objects.filter(
                    pago=True,
                    data_pagamento__gte=data_inicio_dt,
                    data_pagamento__lt=data_final_dt_plus,
                )

                if lojas_qs:
                    parcelas_qs = parcelas_qs.filter(pagamento__loja__in=lojas_qs)

                # Otimizações de consulta para o template
                parcelas_qs = parcelas_qs.select_related(
                    'pagamento',
                    'pagamento__venda',
                    'pagamento__venda__cliente',
                    'pagamento__venda__loja',
                ).order_by('data_pagamento', 'pagamento_id', 'numero_parcela')

                parcelas_list = list(parcelas_qs)

                # Anexa valor pago calculado para uso direto no template
                for p in parcelas_list:
                    p.paid_amount = p.valor_pago_efetivo

                total_pago = sum(p.paid_amount for p in parcelas_list)

                # Monta contexto específico para listagem por parcelas
                context = super().get_context_data(**kwargs)
                context['list_by_parcelas'] = True
                context['parcelas'] = parcelas_list
                context['lojas'] = lojas_qs.values_list('nome', flat=True) if lojas_qs else []
                context['status_list'] = [self.STATUS_LABELS.get(status, status) for status in status_list]
                context['data_inicio'] = data_inicio_dt
                context['data_fim'] = data_final_dt

                # Totais relevantes para a visão por parcelas
                context['total_pago'] = total_pago
                context['quantidade_pagamentos'] = len(parcelas_list)

                # Outros totais não se aplicam na visão por parcela; mantém 0 para não quebrar o rodapé existente
                context['total_atrasado'] = 0
                context['total_a_vencer'] = 0
                context['total_proximo_vencimento'] = 0
                context['total'] = 0
                context['total_quitado'] = 0

                return context

            # Subqueries para datas relevantes
            proximo_vencimento_subquery = Subquery(
                Parcela.objects.filter(
                    pagamento=OuterRef('pk'),
                    pago=False,
                    data_vencimento__gte=timezone.now()
                ).order_by('data_vencimento').values('data_vencimento')[:1],
                output_field=DateField()
            )
            ultimo_vencimento_subquery = Subquery(
                Parcela.objects.filter(
                    pagamento=OuterRef('pk'),
                    pago=False,
                    data_vencimento__lt=timezone.now()
                ).order_by('data_vencimento').values('data_vencimento')[:1],
                output_field=DateField()
            )
            ultimo_pagamento_subquery = Subquery(
                Parcela.objects.filter(
                    pagamento=OuterRef('pk'),
                    pago=True
                ).order_by('data_pagamento').values('data_pagamento')[:1],
                output_field=DateField()
            )

            pagamentos_qs = Pagamento.objects.filter(venda__is_deleted=False,devolucao=False).with_status_flags().distinct()
            pagamentos_qs = pagamentos_qs.annotate(
                proximo_vencimento=proximo_vencimento_subquery,
                ultimo_vencimento=ultimo_vencimento_subquery,
                ultimo_pagamento=ultimo_pagamento_subquery
            )

            if lojas_qs:
                pagamentos_qs = pagamentos_qs.filter(loja__in=lojas_qs)

            if status_pagamento_ids:
                pagamentos_qs = pagamentos_qs.filter(statuses__in=status_pagamento_ids).distinct()

            # Filtro de status e datas
            if status_list and 'todos' not in status_list:
                q_status = None
                date_filter = Q()
                for status in status_list:
                    if status == 'pendente':
                        q = Q(com_pagamento_pendente=True)
                        date_q = Q(proximo_vencimento__isnull=False, proximo_vencimento__gte=data_inicio_dt, proximo_vencimento__lt=data_final_dt_plus)
                    elif status == 'pago':
                        q = Q()  # Não filtra por flag, só pela data do último pagamento
                        date_q = Q(ultimo_pagamento__isnull=False, ultimo_pagamento__gte=data_inicio_dt, ultimo_pagamento__lt=data_final_dt_plus)
                    elif status == 'atrasado':
                        q = Q(com_parcela_atrasada=True)
                        date_q = Q(ultimo_vencimento__isnull=False, ultimo_vencimento__gte=data_inicio_dt, ultimo_vencimento__lt=data_final_dt_plus)
                    elif status == 'desativado':
                        q = Q(desativado=True)
                        date_q = Q()  # Para desativados, não filtra por data
                    elif status == 'bloqueado':
                        q = Q(bloqueado=True)
                        date_q = Q()  # Para bloqueados, não filtra por data
                    elif status == 'sem_contato':
                        q = Q(sem_contato=True)
                        date_q = Q()  # Não filtra por data
                    elif status == "quitado":
                        q = Q(todas_parcelas_pagas=True)
                        date_q = Q()  # Não filtra por data
                    elif status == 'mais_prazo':
                        q = Q(mais_prazo=True)
                        date_q = Q()  # Não filtra por data
                    elif status == 'lembrete':
                        q = Q(lembrete=True)
                        date_q = Q()  # Não filtra por data
                    elif status == 'flag_atrasado':
                        q = Q(flag_atrasado=True)
                        date_q = Q()  # Não filtra por data
                    elif status == 'sem_conexao':
                        q = Q(sem_conexao=True)
                        date_q = Q()  # Não filtra por data
                    elif status == 'roubo':
                        q = Q(roubo=True)
                        date_q = Q()  # Não filtra por data
                    elif status == 'devolucao':
                        q = Q(devolucao=True)
                        date_q = Q()  # Não filtra por data
                    else:
                        continue
                    q_status = q if q_status is None else q_status | q
                    date_filter = date_filter | date_q
                if q_status is not None:
                    pagamentos_qs = pagamentos_qs.filter(q_status)
                if date_filter:
                    pagamentos_qs = pagamentos_qs.filter(date_filter)
            else:
                # Caso 'todos' esteja selecionado, considerar qualquer data relevante no intervalo
                if status_list and 'todos' in status_list:
                    pagamentos_qs = pagamentos_qs.filter(
                        Q(proximo_vencimento__isnull=False, proximo_vencimento__gte=data_inicio_dt, proximo_vencimento__lt=data_final_dt_plus) |
                        Q(ultimo_vencimento__isnull=False, ultimo_vencimento__gte=data_inicio_dt, ultimo_vencimento__lt=data_final_dt_plus) |
                        Q(ultimo_pagamento__isnull=False, ultimo_pagamento__gte=data_inicio_dt, ultimo_pagamento__lt=data_final_dt_plus) |
                        Q(sem_conexao=True)
                    ).distinct()
                else:
                    # Se não filtrar por status, usar proximo_vencimento por padrão
                    pagamentos_qs = pagamentos_qs.filter(
                        proximo_vencimento__isnull=False,
                        proximo_vencimento__gte=data_inicio_dt,
                        proximo_vencimento__lt=data_final_dt_plus
                    )

            contas_a_receber = list(pagamentos_qs)

        total_atrasado = sum(
            pagamento.total_atrasos() for pagamento in contas_a_receber if contas_a_receber
        )
        
        total_pago = sum(
            pagamento.valor_pago_ultimo() for pagamento in contas_a_receber if contas_a_receber
        )

        total_a_vencer = sum(
            pagamento.valor_a_vencer() for pagamento in contas_a_receber if contas_a_receber
        )

        total_proximo_vencimento = sum(
            pagamento.valor_atual_a_vencer() for pagamento in contas_a_receber if contas_a_receber
        )

        total = sum(
            pagamento.valor_total_parcelas() for pagamento in contas_a_receber
        )

        total_quitado = sum(
            pagamento.valor_quitado() for pagamento in contas_a_receber if contas_a_receber
        )

        context = super().get_context_data(**kwargs)
        context['contas_a_receber'] = contas_a_receber
        context['lojas'] = Loja.objects.filter(id__in=lojas).values_list('nome', flat=True)
        context['status_list'] = [self.STATUS_LABELS.get(status, status) for status in status_list]
        context['data_inicio'] = datetime.strptime(data_inicio, "%Y-%m-%d").date() if data_inicio else None
        context['data_fim'] = datetime.strptime(data_fim, "%Y-%m-%d").date() if data_fim else None
        context['total_atrasado'] = total_atrasado
        context['total_pago'] = total_pago
        context['total_a_vencer'] = total_a_vencer
        context['total_proximo_vencimento'] = total_proximo_vencimento
        context['total'] = total
        context['total_quitado'] = total_quitado

        return context


class RelatorioContasAReceberAvancadoView(BaseView, PermissionRequiredMixin, TemplateView):
    template_name = 'contas_a_receber/relatorio_avancado.html'
    permission_required = 'vendas.can_genarate_report_payments'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = RelatorioContasAReceberAvancadoForm()
        return context


class FolhaRelatorioContasAReceberAvancadoView(BaseView, PermissionRequiredMixin, TemplateView):
    template_name = 'contas_a_receber/folha_relatorio_avancado.html'
    permission_required = 'vendas.can_genarate_report_payments'
    SITUACAO_LABELS = dict(RelatorioContasAReceberAvancadoForm.base_fields['situacoes'].choices)

    def get_context_data(self, **kwargs):
        data_inicio = self.request.GET.get('data_inicial')
        data_fim = self.request.GET.get('data_final')
        lojas = self.request.GET.getlist('lojas')
        situacoes = self.request.GET.getlist('situacoes')
        status_pagamento_ids = self.request.GET.getlist('status_pagamento')

        pagamentos_qs = Pagamento.objects.exclude(venda__is_deleted=True).with_status_flags().distinct()

        if lojas:
            pagamentos_qs = pagamentos_qs.filter(loja__in=lojas)

        # Data da venda
        if data_inicio and data_fim:
            di = datetime.strptime(data_inicio, '%Y-%m-%d')
            df = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            pagamentos_qs = pagamentos_qs.filter(venda__criado_em__gte=di, venda__criado_em__lt=df)

        # Filtros de situações (booleans)
        for s in situacoes:
            if s in ['lembrete','bo','flag_atrasado','sem_conexao','roubo','bloqueado','desativado','devolucao','sem_contato','mais_prazo']:
                kwargs_filter = {s: True}
                pagamentos_qs = pagamentos_qs.filter(**kwargs_filter)

        # Filtro por StatusPagamento M2M
        if status_pagamento_ids:
            pagamentos_qs = pagamentos_qs.filter(statuses__in=status_pagamento_ids).distinct()

        # Annotações para data de vencimento (próximo)
        proximo_vencimento_subquery = Subquery(
            Parcela.objects.filter(
                pagamento=OuterRef('pk'),
                pago=False,
                data_vencimento__gte=timezone.now()
            ).order_by('data_vencimento').values('data_vencimento')[:1],
            output_field=DateField()
        )
        pagamentos_qs = pagamentos_qs.annotate(proximo_vencimento=proximo_vencimento_subquery)

        # Carrega relacionamentos necessários
        contatos_qs = Contato.objects.select_related('criado_por').order_by('-criado_em', '-id')
        pagamentos_qs = pagamentos_qs.select_related(
            'venda',
            'venda__cliente',
            'venda__loja',
        ).prefetch_related(
            Prefetch('venda__cliente__contatos', queryset=contatos_qs, to_attr='contatos_ordenados')
        )

        context = super().get_context_data(**kwargs)
        contas = list(pagamentos_qs)
        for pagamento in contas:
            contatos = getattr(pagamento.venda.cliente, 'contatos_ordenados', [])
            pagamento.ultimo_contato = contatos[0] if contatos else None

        context['contas'] = contas
        context['data_inicio'] = data_inicio
        context['data_fim'] = data_fim
        context['situacoes'] = [self.SITUACAO_LABELS.get(situacao, situacao) for situacao in situacoes]
        return context

class RelatorioSaidaView(BaseView, PermissionRequiredMixin, TemplateView):
    template_name = 'relatorio/relatorio_saida.html'
    permission_required = 'vendas.can_generate_report_sale'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = RelatorioSaidaForm()

        return context
    
class FolhaRelatorioSaidaView(BaseView, PermissionRequiredMixin, TemplateView):
    template_name = 'relatorio/folha_relatorio_saida.html'
    permission_required = 'financeiro.view_caixamensal'

    def get_context_data(self, **kwargs):
        data_inicio = self.request.GET.get('data_inicial')
        data_fim = self.request.GET.get('data_final')
        lojas = self.request.GET.getlist('lojas')
        saidas = []

        if data_inicio and data_fim:
            lojas = Loja.objects.filter(id__in=lojas)
            data_final = datetime.strptime(data_fim, "%Y-%m-%d").date() + timedelta(days=1)

            for loja in lojas:
                caixas = Caixa.objects.filter(loja=loja).filter(data_abertura__range=[data_inicio, data_final]).filter(data_fechamento__isnull=False)
                for caixa in caixas:
                    saidas += caixa.lancamentos_caixa.filter(tipo_lancamento='2')

        context = super().get_context_data(**kwargs)
        context['saidas'] = saidas
        # retornar apenas os nomes das lojas
        context['lojas'] = lojas.values_list('nome', flat=True)
        context['data_inicio'] = datetime.strptime(data_inicio, "%Y-%m-%d").date() if data_inicio else None
        context['data_fim'] = datetime.strptime(data_fim, "%Y-%m-%d").date() if data_fim else None
        context['total_saida'] = sum(saida.valor for saida in saidas)
        context['quantidade_saida'] = len(saidas)

        return context
    
    
    

class RepasseCreateView(PermissionRequiredMixin, CreateView):
    model = Repasse
    form_class = RepasseForm
    permission_required = 'financeiro.add_repasse'

    def form_valid(self, form):
        form.instance.criado_por = self.request.user
        form.instance.criado_em = timezone.now()
        form.instance.atualizado_por = self.request.user
        form.instance.atualizado_em = timezone.now()
        messages.success(self.request, "Repasse criado com sucesso!")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('vendas:loja_detail', kwargs={'pk': self.object.loja.pk})
    

class RepasseUpdateView(PermissionRequiredMixin, UpdateView):
    model = Repasse
    form_class = RepasseForm
    template_name = 'repasse/repasse_form.html'
    permission_required = 'financeiro.change_repasse'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['loja'] = self.object.loja
        return context

    def form_valid(self, form):
        form.instance.atualizado_por = self.request.user
        form.instance.atualizado_em = timezone.now()
        messages.success(self.request, "Repasse atualizado com sucesso!")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('vendas:loja_detail', kwargs={'pk': self.object.loja.pk})
    
    
class ContatoCreateView(PermissionRequiredMixin, CreateView):
    model = Contato
    form_class = ContatoForm
    permission_required = 'vendas.add_contato'

    def dispatch(self, request, *args, **kwargs):
        # só POST via modal
        if request.method != 'POST':
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        pagamento = get_object_or_404(Pagamento, pk=self.kwargs['pagamento_pk'])
        form.instance.cliente = pagamento.venda.cliente
        form.instance.criado_por = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('financeiro:contas_a_receber_update', args=[self.kwargs['pagamento_pk']])


class ContatoUpdateView(PermissionRequiredMixin, UpdateView):
    model = Contato
    form_class = ContatoForm
    permission_required = 'vendas.change_contato'

    def form_valid(self, form):
        form.instance.modificado_por = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        # volta à tela de pagamento
        return reverse('financeiro:contas_a_receber_update', args=[self.kwargs['pagamento_pk']])


# ─────────────────────────────────────────────────────────────────────────────
# Relatório de Parcelas — grade colorida por status (Excel)
#   🟢 Pago  •  🔴 Atrasado  •  🟡 A vencer
# ─────────────────────────────────────────────────────────────────────────────

def _brl(valor):
    """Formata número/Decimal como moeda brasileira: 1234.5 -> 'R$ 1.234,50'."""
    s = f"{float(valor or 0):,.2f}"  # 1,234.50
    return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")


class _FiltroParcelasMixin:
    """Monta o queryset de Pagamentos parcelados aplicando os filtros do relatório."""

    def get_pagamentos(self):
        request = self.request
        user = request.user

        qs = (
            Pagamento.objects
            .exclude(venda__is_deleted=True)
            .filter(tipo_pagamento__parcelas=True)
            .exclude(tipo_pagamento__caixa=True)
            .with_status_flags()
        )

        # Escopo por loja (quem não vê todos os pagamentos fica preso à loja da sessão)
        if not user.has_perm('vendas.can_view_all_payments'):
            qs = qs.filter(loja_id=request.session.get('loja_id'))
        else:
            lojas = request.GET.getlist('lojas')
            if lojas:
                qs = qs.filter(loja_id__in=lojas)

        # Status
        status = request.GET.get('status') or 'todos'
        if status == 'atrasado':
            qs = qs.filter(com_parcela_atrasada=True)
        elif status == 'a_vencer':
            qs = qs.filter(com_pagamento_pendente=True)
        elif status == 'quitado':
            qs = qs.filter(todas_parcelas_pagas=True)

        # Período de vencimento das parcelas
        data_inicial = parse_date(request.GET.get('data_inicial') or '')
        data_final = parse_date(request.GET.get('data_final') or '')
        if data_inicial:
            qs = qs.filter(parcelas_pagamento__data_vencimento__gte=data_inicial)
        if data_final:
            qs = qs.filter(parcelas_pagamento__data_vencimento__lte=data_final)

        # Busca por cliente/CPF
        busca = (request.GET.get('busca') or '').strip()
        if busca:
            qs = qs.filter(
                Q(venda__cliente__nome__icontains=busca) |
                Q(venda__cliente__cpf__icontains=busca)
            )

        return (
            qs.distinct()
            .select_related('venda', 'venda__cliente', 'venda__loja', 'loja')
            .prefetch_related('parcelas_pagamento', 'venda__itens_venda__produto')
            .order_by('venda__loja__nome', 'venda__cliente__nome')
        )


class RelatorioParcelasView(BaseView, PermissionRequiredMixin, TemplateView):
    """Página com o formulário de filtros que dispara o download do Excel."""
    template_name = 'contas_a_receber/relatorio_parcelas.html'
    permission_required = 'vendas.can_genarate_report_payments'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = RelatorioParcelasForm(self.request.GET or None)
        return context


class RelatorioParcelasExcelView(_FiltroParcelasMixin, PermissionRequiredMixin, View):
    """Gera o .xlsx em grade de parcelas colorido por status."""
    permission_required = 'vendas.can_genarate_report_payments'

    FILL_HEADER = PatternFill('solid', fgColor='ED7D31')
    FILL_PAGO = PatternFill('solid', fgColor='92D050')
    FILL_ATRASADO = PatternFill('solid', fgColor='FF0000')
    FILL_AVENCER = PatternFill('solid', fgColor='FFC000')
    THIN = Side(style='thin', color='C9A27E')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def get(self, request, *args, **kwargs):
        hoje = timezone.localdate()
        pagamentos = list(self.get_pagamentos())

        # Pré-processa linhas e descobre o maior número de parcelas
        linhas = []
        max_parcelas = 1
        for pag in pagamentos:
            parcelas = sorted(pag.parcelas_pagamento.all(), key=lambda p: p.numero_parcela)
            max_parcelas = max(max_parcelas, len(parcelas))
            venda = pag.venda
            cliente = getattr(venda, 'cliente', None)
            aparelhos = ', '.join(
                i.produto.nome for i in venda.itens_venda.all() if getattr(i, 'produto', None)
            ) if venda else ''
            linhas.append({
                'loja': venda.loja.nome if venda and venda.loja else (pag.loja.nome if pag.loja else ''),
                'cliente': cliente.nome if cliente else '—',
                'aparelho': aparelhos or '—',
                'parcelas': parcelas,
            })

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Parcelamento'

        col_fixas = 3        # Loja, Cliente, Aparelho
        col_totais = 4       # Total, Pago, Atrasado, A vencer
        total_cols = col_fixas + max_parcelas + col_totais

        # ── Título
        titulo = ws.cell(row=1, column=1, value='Relatório de Parcelas')
        titulo.font = Font(bold=True, size=16, color='7A3E00')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        gerado = timezone.localtime().strftime('%d/%m/%Y %H:%M')
        sub = ws.cell(row=2, column=1, value=f'Gerado em {gerado} • {len(linhas)} cliente(s)')
        sub.font = Font(italic=True, color='7A7A7A')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

        # ── Legenda
        ws.cell(row=3, column=1, value='Legenda:').font = Font(bold=True)
        for i, (texto, fill, branco) in enumerate([
            ('Pago', self.FILL_PAGO, False),
            ('Atrasado', self.FILL_ATRASADO, True),
            ('A vencer', self.FILL_AVENCER, False),
        ]):
            cel = ws.cell(row=3, column=2 + i, value=texto)
            cel.fill = fill
            cel.border = self.BORDER
            cel.alignment = Alignment(horizontal='center')
            cel.font = Font(bold=True, color='FFFFFF') if branco else Font(bold=True)

        # ── Cabeçalho
        header_row = 5
        headers = (
            ['Loja', 'Cliente', 'Aparelho']
            + [str(i) for i in range(1, max_parcelas + 1)]
            + ['Total', 'Pago', 'Atrasado', 'A vencer']
        )
        for idx, txt in enumerate(headers, start=1):
            cel = ws.cell(row=header_row, column=idx, value=txt)
            cel.fill = self.FILL_HEADER
            cel.font = Font(bold=True, color='FFFFFF')
            cel.alignment = Alignment(horizontal='center', vertical='center')
            cel.border = self.BORDER

        # ── Dados
        r = header_row + 1
        for linha in linhas:
            ws.cell(row=r, column=1, value=linha['loja']).border = self.BORDER
            ws.cell(row=r, column=2, value=linha['cliente']).border = self.BORDER
            ws.cell(row=r, column=3, value=linha['aparelho']).border = self.BORDER

            tot = tot_pago = tot_atraso = tot_avencer = Decimal('0')
            for p in linha['parcelas']:
                col = col_fixas + p.numero_parcela
                if col > col_fixas + max_parcelas:
                    continue
                valor = p.valor or Decimal('0')
                tot += valor
                if p.pago:
                    fill, font = self.FILL_PAGO, Font(bold=True)
                    tot_pago += p.valor_pago_efetivo
                elif p.data_vencimento and p.data_vencimento < hoje:
                    fill, font = self.FILL_ATRASADO, Font(bold=True, color='FFFFFF')
                    tot_atraso += p.valor_restante
                else:
                    fill, font = self.FILL_AVENCER, Font(bold=True)
                    tot_avencer += p.valor_restante
                venc = p.data_vencimento.strftime('%d/%m/%Y') if p.data_vencimento else ''
                cel = ws.cell(row=r, column=col, value=f'{_brl(valor)}\n{venc}')
                cel.fill = fill
                cel.font = font
                cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cel.border = self.BORDER

            base = col_fixas + max_parcelas
            for offset, valor in enumerate([tot, tot_pago, tot_atraso, tot_avencer], start=1):
                cel = ws.cell(row=r, column=base + offset, value=_brl(valor))
                cel.border = self.BORDER
                cel.alignment = Alignment(horizontal='right')
            r += 1

        # ── Larguras / congelamento / grid
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 26
        for i in range(1, max_parcelas + 1):
            ws.column_dimensions[get_column_letter(col_fixas + i)].width = 14
        for i in range(1, col_totais + 1):
            ws.column_dimensions[get_column_letter(col_fixas + max_parcelas + i)].width = 13
        ws.freeze_panes = ws.cell(row=header_row + 1, column=col_fixas + 1)
        ws.sheet_view.showGridLines = False

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        nome = f'relatorio_parcelas_{hoje.strftime("%Y%m%d")}.xlsx'
        resp = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{nome}"'
        return resp
